
from copy import copy


from openpyxl import load_workbook, Workbook


new_row = 0
last_row = 0


def copySheet(sourceSheet, newSheet):

    global new_row
    global last_row

    for sheet in sourceSheet:
        for row in sheet.worksheets[0].rows:
            new_row = last_row+cell.row
            print("new row:", new_row)
            for cell in row:
                newCell = newSheet.cell(row=new_row, column=cell.col_idx,
                                        value=cell.value)

        new_row = cell.row - 1
# for row in enumerate(sourceSheet2.worksheets[0].iter_rows(min_row=2), start=2):
#     for cell in row[1]:
#         new_row = last_row+cell.row
#         newCell = newSheet.cell(row=new_row, column=cell.col_idx,
#                                 value=cell.value)


filepaths = ['./invoice.xlsx', './invoice2.xlsx']

filepath = [load_workbook(f) for f in filepaths]
merged = Workbook()
o = merged.create_sheet('Sheet 4')
safeTitle = o.title
copySheet(filepath, merged[safeTitle])

merged.save('consolidated.xlsx')
