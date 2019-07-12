# Add sheet to the existing xlsx
from openpyxl import load_workbook
filepath = './demo.xlsx'
wb = load_workbook(filepath)

wb.create_sheet('Sheet2')
wb.save(filepath)