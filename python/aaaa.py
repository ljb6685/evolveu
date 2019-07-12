from copy import copy


from openpyxl import load_workbook, Workbook


def copySheet(sourceSheet, newSheet):
    for sheet in sourceSheet.worksheets[0]:
        for row in sheet.row:
            for cell in row:
                newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                                        value=cell.value)


filepaths = ['./invoice.xlsx', './invoice2.xlsx']

filepath = [load_workbook(f) for f in filepaths]
print(filepath)
merged = Workbook()
o = merged.create_sheet('Sheet 4')
safeTitle = o.title
o2 = merged.create_sheet('Sheet 2')
safeTitle2 = o2.title
copySheet(filepath, merged[safeTitle])

print(merged)
merged.save('consolidated.xlsx')
