# Reading from an Excel sheet - Iterating by rows
from openpyxl import load_workbook
filepath = './demo.xlsx'
wb = load_workbook(filepath)
sheet = wb.active

max_row = sheet.max_row
max_column = sheet.max_column

for i in range(1, max_row+1):
    for j in range(1, max_column+1):
        cell_obj = sheet.cell(row=i, column=j)
        print(cell_obj.value, end=' | ')
    print('\n')