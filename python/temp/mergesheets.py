
from copy import copy
from openpyxl import load_workbook, Workbook


def copySheet(sourceSheet, newSheet):
    counter = 0
    last_row = 0
    for sheet in sourceSheet:

        if(counter == 0):
            for row in sheet.worksheets[0].rows:
                for cell in row:
                    newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                                            value=cell.value)
            last_row = cell.row - 1
            counter = 1
        else:
            row_prev = last_row
            for row in enumerate(sheet.worksheets[0].iter_rows(min_row=2), start=2):
                for cell in row[1]:
                    new_row = last_row+cell.row
                    newCell = newSheet.cell(row=new_row, column=cell.col_idx,
                                            value=cell.value)
            last_row = cell.row + row_prev - 1
            counter = counter + 1


filepaths = ['./invoice.xlsx', './invoice2.xlsx', './invoice.xlsx']
filepath = [load_workbook(f) for f in filepaths]
merged = Workbook()
o = merged.create_sheet('Sheet')
safeTitle = o.title
copySheet(filepath, merged[safeTitle])
merged.save('consolidated.xlsx')
