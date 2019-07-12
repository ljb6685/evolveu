# Reading from an Excel sheet - Reading a cell
from openpyxl import load_workbook
filepath = './demo.xlsx'
wb = load_workbook(filepath)
sheet = wb.active

b1 = sheet['B1'].value
b2 = sheet['B2'].value
b3 = sheet.cell(row=3, column=3).value
print(b1, '\n', b2, '\n', b3)