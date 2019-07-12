# Add data to the Excel sheet - Writing to a cell
from openpyxl import load_workbook
filepath = './demo.xlsx'
wb = load_workbook(filepath)
sheet = wb.active
sheet['A1'] = 1
sheet['B1'] = 'Hello'
sheet.cell(row=3, column=3).value = 5
wb.save(filepath)